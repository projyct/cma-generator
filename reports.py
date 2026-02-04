"""
Report generation module for CMA reports
Supports PDF, Excel, and HTML export formats
"""

from datetime import datetime
from typing import List, Dict
from io import BytesIO
import os


class CMAReportGenerator:
    """Generate CMA reports in various formats"""

    def __init__(self):
        # Lazy-load reportlab only when needed for PDF generation
        self.styles = None
        self._styles_initialized = False

    def _setup_custom_styles(self):
        """Setup custom paragraph styles for PDF"""
        if self._styles_initialized:
            return

        # Import reportlab modules only when needed
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a4d2e'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))

        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2d5f3f'),
            spaceAfter=12,
            spaceBefore=12
        ))

        self.styles.add(ParagraphStyle(
            name='SubjectInfo',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=6
        ))

        self._styles_initialized = True

    def generate_pdf(self, cma_data: Dict, output_path: str):
        """
        Generate PDF CMA report

        Args:
            cma_data: Dictionary containing CMA analysis data
            output_path: Path to save PDF file
        """
        # Lazy-load reportlab modules
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        # Initialize styles if needed
        self._setup_custom_styles()

        doc = SimpleDocTemplate(output_path, pagesize=letter,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)

        story = []

        # Title
        title = Paragraph("Comparative Market Analysis", self.styles['CustomTitle'])
        story.append(title)

        # Company info
        company = Paragraph("CMA Generator", self.styles['Normal'])
        story.append(company)

        report_date = Paragraph(f"Report Date: {datetime.now().strftime('%B %d, %Y')}",
                               self.styles['Normal'])
        story.append(report_date)
        story.append(Spacer(1, 0.3*inch))

        # Subject Property Information
        story.append(Paragraph("Subject Property", self.styles['CustomHeading']))

        subject_info = [
            f"<b>Address:</b> {cma_data.get('subject_address', 'N/A')}",
            f"<b>Bedrooms:</b> {cma_data.get('subject_beds', 'N/A')} | "
            f"<b>Bathrooms:</b> {cma_data.get('subject_baths', 'N/A')} | "
            f"<b>Square Feet:</b> {cma_data.get('subject_sqft', 'N/A'):,}" if cma_data.get('subject_sqft') else "",
        ]

        for info in subject_info:
            if info:
                story.append(Paragraph(info, self.styles['SubjectInfo']))

        story.append(Spacer(1, 0.3*inch))

        # Rent Analysis Summary
        story.append(Paragraph("Rent Analysis Summary", self.styles['CustomHeading']))

        rent_stats = cma_data.get('rent_stats', {})
        stats_data = [
            ['Metric', 'Value'],
            ['Average Rent', f"${rent_stats.get('avg_rent', 0):,.2f}"],
            ['Median Rent', f"${rent_stats.get('median_rent', 0):,.2f}"],
            ['Minimum Rent', f"${rent_stats.get('min_rent', 0):,.2f}"],
            ['Maximum Rent', f"${rent_stats.get('max_rent', 0):,.2f}"],
            ['Suggested Range', f"${rent_stats.get('suggested_low', 0):,.2f} - ${rent_stats.get('suggested_high', 0):,.2f}"],
            ['Number of Comparables', str(rent_stats.get('comp_count', 0))]
        ]

        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a4d2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(stats_table)
        story.append(Spacer(1, 0.3*inch))

        # Comparables Table
        story.append(Paragraph("Comparable Properties", self.styles['CustomHeading']))

        comps = cma_data.get('comparables', [])
        if comps:
            comp_data = [['Address', 'Beds/Baths', 'Sqft', 'Rent', 'Source', 'Distance']]

            for comp in comps[:10]:  # Limit to 10 comps on first page
                # Use display_address if available, fallback to address
                display_addr = comp.get('display_address', comp.get('address', 'N/A'))
                comp_data.append([
                    display_addr[:40],
                    f"{comp.get('bedrooms', 'N/A')}/{comp.get('bathrooms', 'N/A')}",
                    f"{comp.get('sqft', 'N/A'):,}" if comp.get('sqft') else 'N/A',
                    f"${comp.get('rent', 0):,.2f}",
                    comp.get('source', 'Internal'),
                    f"{comp.get('distance_miles', 0):.2f} mi"
                ])

            comp_table = Table(comp_data, colWidths=[2.2*inch, 0.8*inch, 0.7*inch, 0.9*inch, 0.8*inch, 0.8*inch])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d5f3f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9)
            ]))

            story.append(comp_table)

        # Disclaimer
        story.append(Spacer(1, 0.5*inch))
        disclaimer = Paragraph(
            "<i>This Comparative Market Analysis is provided for informational purposes only and should not be "
            "considered a formal appraisal. Actual rental rates may vary based on market conditions, property "
            "condition, and other factors.</i>",
            self.styles['Normal']
        )
        story.append(disclaimer)

        # Build PDF
        doc.build(story)

    def generate_excel(self, cma_data: Dict, output_path: str):
        """
        Generate Excel CMA report with charts

        Args:
            cma_data: Dictionary containing CMA analysis data
            output_path: Path to save Excel file
        """
        # Lazy-load openpyxl modules
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header
        ws_summary['A1'] = "Comparative Market Analysis"
        ws_summary['A1'].font = Font(size=18, bold=True, color="1a4d2e")
        ws_summary['A2'] = "CMA Generator"
        ws_summary['A3'] = f"Report Date: {datetime.now().strftime('%B %d, %Y')}"

        # Subject Property
        ws_summary['A5'] = "Subject Property"
        ws_summary['A5'].font = Font(size=14, bold=True)
        ws_summary['A6'] = "Address:"
        ws_summary['B6'] = cma_data.get('subject_address', 'N/A')
        ws_summary['A7'] = "Bedrooms:"
        ws_summary['B7'] = cma_data.get('subject_beds', 'N/A')
        ws_summary['A8'] = "Bathrooms:"
        ws_summary['B8'] = cma_data.get('subject_baths', 'N/A')
        ws_summary['A9'] = "Square Feet:"
        ws_summary['B9'] = cma_data.get('subject_sqft', 'N/A')

        # Rent Analysis
        rent_stats = cma_data.get('rent_stats', {})
        ws_summary['A11'] = "Rent Analysis"
        ws_summary['A11'].font = Font(size=14, bold=True)

        stats_rows = [
            ('Average Rent', rent_stats.get('avg_rent', 0)),
            ('Median Rent', rent_stats.get('median_rent', 0)),
            ('Minimum Rent', rent_stats.get('min_rent', 0)),
            ('Maximum Rent', rent_stats.get('max_rent', 0)),
            ('Suggested Low', rent_stats.get('suggested_low', 0)),
            ('Suggested High', rent_stats.get('suggested_high', 0)),
            ('Number of Comps', rent_stats.get('comp_count', 0))
        ]

        for idx, (label, value) in enumerate(stats_rows, start=12):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value if label == 'Number of Comps' else f"${value:,.2f}"

        # Comparables Sheet
        ws_comps = wb.create_sheet("Comparables")

        headers = ['Address', 'Bedrooms', 'Bathrooms', 'Sqft', 'Rent', 'Source', 'Status', 'Distance (mi)', 'Notes']
        ws_comps.append(headers)

        # Style header
        header_fill = PatternFill(start_color="2d5f3f", end_color="2d5f3f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws_comps[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add comparable data
        comps = cma_data.get('comparables', [])
        for comp in comps:
            ws_comps.append([
                comp.get('display_address', comp.get('address', 'N/A')),
                comp.get('bedrooms', ''),
                comp.get('bathrooms', ''),
                comp.get('sqft', ''),
                comp.get('rent', 0),
                comp.get('source', 'Internal'),
                comp.get('status', ''),
                round(comp.get('distance_miles', 0), 2),
                comp.get('notes', '')
            ])

        # Auto-adjust column widths
        for column in ws_comps.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_comps.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)

    def generate_html(self, cma_data: Dict, output_path: str, map_html: str = None):
        """
        Generate HTML CMA report with simplified formatting for Google Docs conversion

        Args:
            cma_data: Dictionary containing CMA analysis data
            output_path: Path to save HTML file
            map_html: Optional HTML string for embedded map (skipped for Google Docs)
        """
        rent_stats = cma_data.get('rent_stats', {})
        comps = cma_data.get('comparables', [])

        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>CMA Report - {cma_data.get('subject_address', 'Property')}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 40px;
            line-height: 1.6;
        }}
        h1 {{
            color: #1a4d2e;
            border-bottom: 2px solid #1a4d2e;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #2d5f3f;
            margin-top: 25px;
            margin-bottom: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            margin-bottom: 20px;
        }}
        th {{
            background-color: #2d5f3f;
            color: white;
            padding: 10px;
            text-align: left;
            border: 1px solid #2d5f3f;
        }}
        td {{
            padding: 8px;
            border: 1px solid #ddd;
        }}
        .info-row td:first-child {{
            font-weight: bold;
            width: 200px;
        }}
    </style>
</head>
<body>
    <h1>Comparative Market Analysis</h1>
    <p><strong>Report Date:</strong> {datetime.now().strftime('%B %d, %Y')}</p>

    <h2>Subject Property</h2>
    <p>
        <strong>Address:</strong> {cma_data.get('subject_address', 'N/A')}<br>
        <strong>Bedrooms:</strong> {cma_data.get('subject_beds', 'N/A')} |
        <strong>Bathrooms:</strong> {cma_data.get('subject_baths', 'N/A')} |
        <strong>Square Feet:</strong> {f"{cma_data.get('subject_sqft', 0):,}" if cma_data.get('subject_sqft') else 'N/A'}
    </p>

    <h2>Rent Analysis Summary</h2>
    <table>
        <tr class="info-row">
            <td>Average Rent</td>
            <td>${rent_stats.get('avg_rent', 0):,.2f}</td>
        </tr>
        <tr class="info-row">
            <td>Median Rent</td>
            <td>${rent_stats.get('median_rent', 0):,.2f}</td>
        </tr>
        <tr class="info-row">
            <td>Suggested Rent Range</td>
            <td>${rent_stats.get('suggested_low', 0):,.2f} - ${rent_stats.get('suggested_high', 0):,.2f}</td>
        </tr>
        <tr class="info-row">
            <td>Number of Comparables</td>
            <td>{rent_stats.get('comp_count', 0)}</td>
        </tr>
    </table>

    <h2>Comparable Properties</h2>
    <table>
        <thead>
            <tr>
                <th>Address</th>
                <th>Beds/Baths</th>
                <th>Sqft</th>
                <th>Rent</th>
                <th>Source</th>
                <th>Distance</th>
            </tr>
        </thead>
        <tbody>
"""

        # Limit to top 10 comparables for cleaner report
        for comp in comps[:10]:
            html_content += f"""
            <tr>
                <td>{comp.get('address', 'N/A')}</td>
                <td>{comp.get('bedrooms', 'N/A')}/{comp.get('bathrooms', 'N/A')}</td>
                <td>{f"{comp.get('sqft', 0):,}" if comp.get('sqft') else 'N/A'}</td>
                <td>${comp.get('rent', 0):,.2f}</td>
                <td>{comp.get('source', 'Internal')}</td>
                <td>{comp.get('distance_miles', 0):.2f} mi</td>
            </tr>
"""

        html_content += """
        </tbody>
    </table>

    <p><em>This Comparative Market Analysis is provided for informational purposes only and should not be considered a formal appraisal. Actual rental rates may vary based on market conditions, property condition, and other factors.</em></p>
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def create_map(self, subject_property: Dict, comparables: List[Dict]) -> str:
        """
        Create an interactive map with subject property and comparables

        Args:
            subject_property: Subject property with lat/lon
            comparables: List of comparable properties with lat/lon

        Returns:
            HTML string of the map
        """
        # Lazy-load folium
        import folium

        # Center map on subject property
        center_lat = subject_property.get('latitude', 35.9940)
        center_lon = subject_property.get('longitude', -78.8986)

        m = folium.Map(location=[center_lat, center_lon], zoom_start=13)

        # Add subject property marker
        folium.Marker(
            [center_lat, center_lon],
            popup=f"<b>SUBJECT</b><br>{subject_property.get('address', 'Subject Property')}",
            icon=folium.Icon(color='red', icon='home', prefix='fa')
        ).add_to(m)

        # Add comparable markers
        for comp in comparables:
            if comp.get('latitude') and comp.get('longitude'):
                popup_text = f"""
                <b>{comp.get('address', 'N/A')}</b><br>
                {comp.get('bedrooms', 'N/A')} bed / {comp.get('bathrooms', 'N/A')} bath<br>
                {comp.get('sqft', 'N/A')} sqft<br>
                Rent: ${comp.get('rent', 0):,.2f}<br>
                Distance: {comp.get('distance_miles', 0):.2f} mi
                """

                folium.Marker(
                    [comp['latitude'], comp['longitude']],
                    popup=popup_text,
                    icon=folium.Icon(color='blue', icon='building', prefix='fa')
                ).add_to(m)

        return m._repr_html_()
